import csv, sqlite3, time
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import date, datetime, timedelta
from pathlib import Path
from zoneinfo import ZoneInfo
import requests
from openpyxl import Workbook

START=date(2021,8,24); END=date(2026,8,24); TZ=ZoneInfo('Europe/Athens')
API='https://api.opap.gr/draws/v3.0/1100/draw-date/{d}/{d}?size=300'
OUT=Path('kino-monitor/export_5years_output'); OUT.mkdir(parents=True,exist_ok=True)
DB=OUT/'KINO_5years_2021-08-24_to_2026-08-24.sqlite3'
CSV=OUT/'KINO_5years_2021-08-24_to_2026-08-24.csv'
XLSX=OUT/'KINO_5years_2021-08-24_to_2026-08-24.xlsx'
SUMMARY=OUT/'KINO_5years_summary.txt'

def days():
 d=START
 while d<=END:
  yield d; d+=timedelta(days=1)

def fetch(d):
 u=API.format(d=d.isoformat()); err=None
 for i in range(7):
  try:
   r=requests.get(u,headers={'Accept':'application/json','User-Agent':'KINO-history-export'},timeout=30)
   if r.status_code==429: time.sleep(5+i*2); continue
   r.raise_for_status(); data=r.json(); rows=[]
   for o in data.get('content',[]):
    nums=[int(x) for x in o.get('winningNumbers',{}).get('list',[])]
    if len(nums)!=20: continue
    ms=int(o['drawTime']); dt=datetime.fromtimestamp(ms/1000,TZ)
    odd=sum(n%2 for n in nums); even=20-odd
    res='ΜΟΝΑ' if odd>even else ('ΖΥΓΑ' if even>odd else 'ΙΣΟΠΑΛΙΑ')
    rows.append((int(o['drawId']),ms,dt.isoformat(timespec='seconds'),dt.date().isoformat(),dt.strftime('%H:%M:%S'),odd,even,res,*nums))
   return d,rows
  except Exception as e:
   err=e; time.sleep(min(20,2**i))
 raise RuntimeError(f'{d}: {err}')

def main():
 if DB.exists(): DB.unlink()
 con=sqlite3.connect(DB)
 nums=','.join(f'n{i} integer not null' for i in range(1,21))
 con.executescript(f'''create table draws(draw_id integer primary key,draw_time_ms integer not null,draw_time_local text not null,draw_date text not null,draw_clock text not null,odd_count integer not null,even_count integer not null,result text not null,{nums},no_odd_streak integer,no_even_streak integer,tie_streak integer);create index idx_time on draws(draw_time_ms);create table metadata(key text primary key,value text not null);''')
 ds=list(days()); zero=[]; failures=[]; total=0
 with ThreadPoolExecutor(max_workers=8) as ex:
  fs={ex.submit(fetch,d):d for d in ds}
  for k,f in enumerate(as_completed(fs),1):
   d=fs[f]
   try: _,rows=f.result()
   except Exception as e: failures.append(str(e)); continue
   if not rows: zero.append(d.isoformat())
   con.executemany('insert or replace into draws('+','.join(['draw_id','draw_time_ms','draw_time_local','draw_date','draw_clock','odd_count','even_count','result']+[f'n{i}' for i in range(1,21)])+') values('+','.join(['?']*28)+')',rows)
   con.commit(); total+=len(rows)
   if k%25==0: print(f'{k}/{len(ds)} days, {total} rows',flush=True)
 if failures: raise RuntimeError('API failures: '+str(failures[:20]))
 no_odd=no_even=ties=0; batch=[]
 for did,res in con.execute('select draw_id,result from draws order by draw_time_ms,draw_id'):
  no_odd=0 if res=='ΜΟΝΑ' else no_odd+1; no_even=0 if res=='ΖΥΓΑ' else no_even+1; ties=ties+1 if res=='ΙΣΟΠΑΛΙΑ' else 0
  batch.append((no_odd,no_even,ties,did))
  if len(batch)>=10000:
   con.executemany('update draws set no_odd_streak=?,no_even_streak=?,tie_streak=? where draw_id=?',batch); batch=[]
 if batch: con.executemany('update draws set no_odd_streak=?,no_even_streak=?,tie_streak=? where draw_id=?',batch)
 con.commit()
 hdr=['draw_id','draw_time_local','draw_date','draw_clock','odd_count','even_count','result']+[f'n{i}' for i in range(1,21)]+['no_odd_streak','no_even_streak','tie_streak']
 q='select '+','.join(hdr)+' from draws order by draw_time_ms,draw_id'
 with CSV.open('w',encoding='utf-8-sig',newline='') as f:
  w=csv.writer(f); w.writerow(hdr); w.writerows(con.execute(q))
 wb=Workbook(write_only=True); ws=wb.create_sheet('KINO draws'); ws.append(hdr)
 for row in con.execute(q): ws.append(row)
 meta=wb.create_sheet('Summary')
 count=con.execute('select count(*) from draws').fetchone()[0]
 first=con.execute('select draw_id,draw_time_local from draws order by draw_time_ms limit 1').fetchone(); last=con.execute('select draw_id,draw_time_local from draws order by draw_time_ms desc limit 1').fetchone()
 rc=dict(con.execute('select result,count(*) from draws group by result'))
 summary={'period':f'{START} to {END}','generated_at_athens':datetime.now(TZ).isoformat(timespec='seconds'),'calendar_days':len(ds),'zero_draw_dates':len(zero),'zero_dates':','.join(zero),'total_unique_draws':count,'first_draw':first,'last_draw':last,'MONA':rc.get('ΜΟΝΑ',0),'ZYGA':rc.get('ΖΥΓΑ',0),'ISOPALIA':rc.get('ΙΣΟΠΑΛΙΑ',0),'source':'Official OPAP API game 1100'}
 meta.append(['Field','Value'])
 for a,b in summary.items(): meta.append([a,str(b)]); con.execute('insert or replace into metadata values(?,?)',(a,str(b)))
 con.commit(); wb.save(XLSX); con.close()
 SUMMARY.write_text('\n'.join(f'{a}: {b}' for a,b in summary.items()),encoding='utf-8')
 print('EXPORT COMPLETE',flush=True)
 for p in (DB,CSV,XLSX,SUMMARY): print(p,p.stat().st_size,flush=True)
if __name__=='__main__': main()
